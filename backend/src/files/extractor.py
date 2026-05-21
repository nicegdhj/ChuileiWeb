import logging
import os

logger = logging.getLogger(__name__)


def _ext(name: str) -> str:
    return os.path.splitext(name)[1].lower()


def extract_text(path: str, mime: str, original_name: str) -> str:
    ext = _ext(original_name)
    try:
        if ext in (".txt", ".md", ".csv", ".log") or mime.startswith("text/"):
            with open(path, "r", encoding="utf-8", errors="replace") as f:
                return f.read()
        if ext == ".pdf" or mime == "application/pdf":
            return _extract_pdf(path)
        if ext == ".docx":
            return _extract_docx(path)
        if ext in (".xlsx", ".xlsm"):
            return _extract_xlsx(path)
    except Exception:
        logger.exception("extract failed for %s", original_name)
    return ""


def _extract_pdf(path: str) -> str:
    from pypdf import PdfReader
    reader = PdfReader(path)
    return "\n".join((page.extract_text() or "") for page in reader.pages)


def _extract_docx(path: str) -> str:
    from docx import Document
    doc = Document(path)
    return "\n".join(p.text for p in doc.paragraphs)


def _extract_xlsx(path: str) -> str:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    rows: list[str] = []
    for ws in wb.worksheets:
        rows.append(f"[Sheet: {ws.title}]")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(c.strip() for c in cells):
                rows.append("\t".join(cells))
    wb.close()
    return "\n".join(rows)
